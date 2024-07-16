# -*- coding: UTF-8 -*-
import os
import shutil
import tkinter as tk
from tkinter import filedialog
from doc2docx import convert
from docx import Document
from openpyxl import load_workbook
from copy import copy

class FileMoverGUI:
    def __init__(self, master):
        """ param used """
        self.source_default = "C:\\Users\\user\\Desktop"  # 源文件夹路径
        self.target_default = "D:\\Documents\\excel.xlsx"  # 目标文件夹路径
        self.file_path = None
        self.mid_dir_path = None
        self.already_done_path = None
        self.jiushiwan_num = 1

        """ window """
        self.master = master  # 主窗口
        self.master.title("定制版")  # 窗口标题
        self.master.geometry("600x600")  # 窗口大小

        # 创建一个标签，显示源文件路径
        self.source_label = tk.Label(self.master, text="源文件路径：")
        self.source_label.grid(row=0, column=0, sticky=tk.W)

        # 创建一个文本框，显示源文件路径
        self.source_entry = tk.Entry(self.master, width=58)
        self.source_entry.grid(row=0, column=1, sticky=tk.W)

        # 使用默认的源文件夹路径来填充文本框
        self.source_entry.insert(0, self.source_default)

        # 创建一个按钮，点击后弹出文件选择对话框
        self.source_button = tk.Button(self.master, text="选择", command=self.select_source_file)
        self.source_button.grid(row=0, column=2, sticky=tk.NSEW)

        # 创建一个标签，显示目标文件夹路径
        self.target_label = tk.Label(self.master, text="目标文件夹路径：")
        self.target_label.grid(row=1, column=0, sticky=tk.W)

        # 创建一个文本框，显示目标文件夹路径
        self.target_entry = tk.Entry(self.master, width=58)
        self.target_entry.grid(row=1, column=1, sticky=tk.W)

        # 使用默认的目标文件夹路径来填充文本框
        self.target_entry.insert(0, self.target_default)

        # 点击后弹出文件夹选择对话框
        self.target_button = tk.Button(self.master, text="选择", command=self.select_target_file)
        self.target_button.grid(row=1, column=2, sticky=tk.NSEW)

        # 标签，显示新文件名
        # self.new_name_label = tk.Label(self.master, text="新文件名：")
        # self.new_name_label.grid(row=2, column=0, sticky=tk.W)
        #
        # # 文本框，输入新文件名
        # self.new_name_entry = tk.Entry(self.master, width=58)
        # self.new_name_entry.grid(row=2, column=1, sticky=tk.W)

        """ 按钮，执行主操作 """
        self.move_button = tk.Button(self.master, text="执行", command=self.runner)
        self.move_button.grid(row=3, column=0, sticky=tk.NSEW)

        self.move_button = tk.Button(self.master, text="没事点着玩！", command=self.jiushiwan)
        self.move_button.grid(row=3, column=1, sticky=tk.NSEW)

        # 创建一个文本区域，显示操作结果信息
        self.info_text = tk.Text(self.master)
        self.info_text.grid(row=4, columnspan=3)

        # 再创建一个滚动条控件，并关联文本区域的yview方法
        self.scroll = tk.Scrollbar(self.master)
        self.scroll.config(command=self.info_text.yview)
        self.info_text.config(yscrollcommand=self.scroll.set)  # 这一行也可以放在创建文本区域控件之后

        # 设置不同颜色的标签
        self.info_text.tag_config("success", foreground="green")
        self.info_text.tag_config("error", foreground="red")

        # 创建一个按钮，点击后清除信息框内容
        self.clear_button = tk.Button(self.master, text="  清除  ", command=self.clear_info_text)
        self.clear_button.grid(row=5, column=0, sticky=tk.NSEW)

    def select_source_file(self):
        # 弹出文件选择对话框，获取源文件路径
        source_folder = filedialog.askdirectory(initialdir=self.source_default)  # 使用默认的源文件夹路径作为初始目录
        # 将源文件路径显示在文本框中
        self.source_entry.delete(0, tk.END)
        self.source_entry.insert(0, source_folder)

        # 更新默认的源文件夹路径为当前选择的源文件所在的目录
        self.source_default = source_folder

    def select_target_file(self):
        # 弹出文件夹选择对话框，获取目标文件夹路径
        target_file = filedialog.askopenfilename(initialdir=self.target_default)  # 使用默认的目标文件夹路径作为初始目录
        # 将目标文件夹路径显示在文本框中
        self.target_entry.delete(0, tk.END)
        self.target_entry.insert(0, target_file)

        # 更新默认的目标文件夹路径为当前选择的目标文件夹
        self.target_default = target_file

    def move_and_rename(self):
        # 获取源文件路径，目标文件夹路径，新文件名
        source_path = self.source_entry.get()
        target_path = self.target_entry.get()
        new_name = self.new_name_entry.get()

        # 判断是否输入了有效的参数
        if source_path and target_path and new_name:
            try:
                # 获取源文件的扩展名
                extension = os.path.splitext(source_path)[1]
                # 拼接目标文件的完整路径和新文件名
                target_file = os.path.join(target_path, new_name + extension)
                # 复制源文件到目标文件
                shutil.copy(source_path, target_file)
                # 删除源文件
                os.remove(source_path)
                # 在文本区域显示成功信息，使用绿色标签
                self.info_text.insert(tk.END, "已经成功移动和重命名了文件\n", "success")
            except Exception as e:
                # 在文本区域显示错误信息，使用红色标签
                self.info_text.insert(tk.END, "移动和重命名文件时发生了错误：" + str(e) + "\n", "error")
        else:
            # 在文本区域显示提示信息，使用红色标签
            self.info_text.insert(tk.END, "请先选择源文件，目标文件夹，和输入新文件名\n", "error")

    def clear_info_text(self):
        self.info_text.delete(1.0, tk.END)

    def operate_mid_dir(self):
        self.mid_dir_path = self.source_default + "\\mid_dir"
        self.already_done_path = self.source_default + "\\AlreadyDone"
        if not os.path.isdir(self.mid_dir_path):
            os.mkdir(self.mid_dir_path)

        if not os.path.isdir(self.already_done_path):
            os.mkdir(self.already_done_path)

    def delet_mid_dir(self):
        shutil.rmtree(self.mid_dir_path)

    def doc2docx(self, file_path):
        convert(file_path, self.mid_dir_path)
        # self.file_path = self.file_path + "x"

    def read_from_word(self, doc_path):
        doc = Document(doc_path)
        # 获取Word文档中的第一个表格
        table = doc.tables[0]
        # 提取表格数据
        table_data = [[cell.text for cell in row.cells] for row in table.rows]

        work_order_number = table_data[0][1]
        signing_time = table_data[3][-1].split()[0]
        classification = table_data[4][-1]
        appeal_content = table_data[5][-1]
        deadline = table_data[10][-1].split()[0]

        if "薪资" in classification:
            classification = "欠薪"
        elif "质量" in classification:
            classification = "质量"
        else:
            classification = '0'

        data = (work_order_number, classification, signing_time, deadline, appeal_content)
        return data

    def get_excel_info(self):
        wb = load_workbook(self.target_default)
        ws = wb.active
        current_index = None
        row_number = None

        for i, row in enumerate(ws.iter_rows(), start=1):
            if row[1].value is not None:
                current_index = row[0].value
                row_number = i

        return current_index + 1, row_number + 1

    def write_to_excel(self, row, value):
        column = 1

        wb = load_workbook(self.target_default)
        ws = wb.active
        for _ in value:
            ws.cell(row=row, column=column, value=value[column - 1])
            column = column + 1

        # 获取上一行单元格的样式
        prev_row = ws[row - 1]

        # 复制上一行的样式到新的行
        for prev_cell in prev_row:
            new_cell = ws.cell(row=row, column=prev_cell.column)
            if new_cell._style is None:
                new_cell._style = copy(prev_cell._style)
            else:
                new_cell.font = copy(prev_cell.font)
                new_cell.border = copy(prev_cell.border)
                new_cell.fill = copy(prev_cell.fill)
                new_cell.number_format = copy(prev_cell.number_format)
                new_cell.protection = copy(prev_cell.protection)
                new_cell.alignment = copy(prev_cell.alignment)

        wb.save(self.target_default)

    def read_and_insert(self):
        self.operate_mid_dir()
        file_names = os.listdir(self.source_default)
        current_index, current_row = self.get_excel_info()

        for file in file_names:

            file_path = self.source_default + "\\" + file
            suffix = os.path.splitext(file_path)[-1]
            if suffix != ".doc" and suffix != ".docx":
                continue

            # start info
            self.info_text.insert(tk.END, file, "success")
            self.info_text.update_idletasks()

            if suffix == ".doc":
                self.info_text.insert(tk.END, "\t 正在转换文件", "success")
                self.info_text.update_idletasks()
                self.doc2docx(file_path)
                self.info_text.insert(tk.END, "\t 转换成功！\n", "success")
                self.info_text.update_idletasks()
                file_path = self.mid_dir_path + "\\" + file + "x"
            data = self.read_from_word(file_path)
            data = (current_index, ) + data

            try:
                self.write_to_excel(current_row, data)
                self.info_text.insert(tk.END, f"\t 写入成功！行号{current_row}，序号{current_index} \n", "success")
                self.info_text.update_idletasks()
                current_row += 1
                current_index += 1
            except Exception as e:
                self.info_text.insert(tk.END, "\t 发生错误！\n" + str(e) + "\n", "error")
                self.info_text.update_idletasks()

            shutil.move(self.source_default + "\\" + file, self.already_done_path)

        self.delet_mid_dir()
        self.info_text.insert(tk.END, "\n\n\n\t -------团------- 大功告成！-------团-------\n", "success")

    def runner(self):
        if not os.path.exists(self.target_default):
            self.info_text.insert(tk.END, "不存在目标excel文件！阿团检查路径\n", "error")
            self.info_text.update_idletasks()
        elif not os.path.isdir(self.source_default):
            self.info_text.insert(tk.END, "不存在源文件夹！阿团检查路径\n", "error")
            self.info_text.update_idletasks()
        else:
            self.read_and_insert()

    def jiushiwan(self):
        self.info_text.insert(tk.END, f"金钱+{self.jiushiwan_num}!\n", "success")
        self.jiushiwan_num += 1
        if self.jiushiwan_num % 10 == 0:
            char = """
               /)/)   (\\(\\
              ( . .)  (‘.’ )
              ( v v ) (' ')(' ')
              (,,)-(,,)-(,,)-(,,)
              """
            self.info_text.insert(tk.END, char)


# 创建一个主窗口
root = tk.Tk()
# 创建一个FileMoverGUI对象
fm_gui = FileMoverGUI(root)
# 进入主循环
root.mainloop()
